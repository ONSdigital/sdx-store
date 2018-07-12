import logging
import time

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from structlog import wrap_logger

logger = wrap_logger(logging.getLogger(__name__))


def create_comments_book(survey_id, comments):
    logger.info("Generating excel file")
    wb = Workbook()
    row = 2
    ws = wb.active
    ws.cell(1, 1, "Survey ID: {}".format(survey_id))
    ws.cell(1, 2, "Comments found: {}".format(str(len(comments))))
    for comment in comments:
        row += 1
        ws.cell(row, 1, comment.data['data']['146'])
        ws.cell(row, 2, comment.data['submitted_at'])

    filename = "comments_{}.xlsx".format(time.strftime("%Y%m%d-%H%M%S"))
    logger.info("Created excel file", filename=filename)
    return save_virtual_workbook(wb), "comments_{}.xlsx".format(time.strftime("%Y%m%d-%H%M%S"))
