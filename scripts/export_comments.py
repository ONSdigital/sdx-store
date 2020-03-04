import os
from string import ascii_lowercase
import sys
parent_dir_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.append(parent_dir_path)

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

from app.models import SurveyResponse
import settings

try:
    db = create_engine(settings.DB_URI)
    Session = sessionmaker(db)
    session = Session()
    base = declarative_base()
except SQLAlchemyError as e:
    print(e)
    raise


def create_comments_excel_file(survey_id, period, submissions):
    """Extract comments from submissions and write them to an excel file"""
    print("Generating Excel file")
    workbook = Workbook()
    row = 2
    surveys_with_comments_count = 0
    ws = workbook.active

    for submission in submissions:
        comment = get_comment_text(submission)

        boxes_selected = ""
        for key in ('146' + letter for letter in ascii_lowercase[0:]):
            if key in submission.data['data'].keys():
                boxes_selected = boxes_selected + key + ' '

        if not comment:
            continue
        row += 1
        surveys_with_comments_count += 1
        ws.cell(row, 1, submission.data['metadata']['ru_ref'])
        ws.cell(row, 2, submission.data['collection']['period'])
        if survey_id == '134':
            if '300w' in submission.data['data']:
                ws.cell(row, 5, submission.data['data']['300w'])
            if '300f' in submission.data['data']:
                ws.cell(row, 6, submission.data['data']['300f'])
            if '300m' in submission.data['data']:
                ws.cell(row, 7, submission.data['data']['300m'])
            if '300w4' in submission.data['data']:
                ws.cell(row, 8, submission.data['data']['300w4'])
            if '300w5' in submission.data['data']:
                ws.cell(row, 9, submission.data['data']['300w5'])
            checkboxes = ['91w', '92w1', '92w2', '94w1', '94w2', '95w', '96w', '97w',
                          '91f', '92f1', '92f2', '94f1', '94f2', '95f', '96f', '97f',
                          '191m', '192m1', '192m2', '194m1', '194m2', '195m', '196m', '197m',
                          '191w4', '192w41', '192w42', '194w41', '194w42', '195w4', '196w4', '197w4',
                          '191w5', '192w51', '192w52', '194w51', '194w52', '195w5', '196w5', '197w5']
            for checkbox in checkboxes:
                if checkbox in submission.data['data']:
                    boxes_selected = boxes_selected + f"{checkbox}, "
        ws.cell(row, 3, boxes_selected)
        ws.cell(row, 4, comment)


    ws.cell(1, 1, f"Survey ID: {survey_id}")
    ws.cell(1, 2, f"Comments found: {surveys_with_comments_count}")
    if survey_id == '134':
        ws.cell(1, 5, "Weekly comment")
        ws.cell(1, 6, "Fortnightly comment")
        ws.cell(1, 7, "Calendar Monthly comment")
        ws.cell(1, 8, "4 Weekly Pay comment")
        ws.cell(1, 9, "5 Weekly Pay comment")
    print(f"{surveys_with_comments_count} out of {len(submissions)} submissions had comments")

    parent_dir_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    filename = os.path.join(parent_dir_path, f"{survey_id}_{period}.xlsx")
    workbook.save(filename)
    workbook.close()
    print(f"Excel file {filename} generated")


def get_all_submissions(survey_id, period):
    """Get all submissions that match the survey_id and period supplied"""
    if survey_id == '181':
        vacancies_records = []
        for survey_id in ['182', '183', '184', '185']:
            survey_records = session.query(SurveyResponse).filter(SurveyResponse.data['survey_id'].astext == survey_id)
            period_specific_records = survey_records.filter(SurveyResponse.data['collection']['period'].astext == period).all()
            vacancies_records = vacancies_records + period_specific_records
            print(f"Retrieved {len(period_specific_records)} submissions for {survey_id}")
        return vacancies_records
    survey_records = session.query(SurveyResponse).filter(SurveyResponse.data['survey_id'].astext == survey_id)
    records = survey_records.filter(SurveyResponse.data['collection']['period'].astext == period).all()
    print(f"Retrieved {len(records)} submissions for {survey_id}")
    return records


def get_comment_text(submission):
    """Returns the respondent typed text from a submission.  The qcode for this text will be different depending
    on the survey
    """
    if submission.data['survey_id'] == '009':
        return submission.data['data'].get('146h')
    if submission.data['survey_id'] == '187':
        return submission.data['data'].get('500')
    if submission.data['survey_id'] == '134':
        return submission.data['data'].get('300')
    return submission.data['data'].get('146')


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("Either survey id or period is missing")

    survey_id = sys.argv[1]
    period = sys.argv[2]

    print(f'Survey id is {survey_id}')
    print(f'Period is {period}')

    try:
        submissions = get_all_submissions(survey_id, period)
    except SQLAlchemyError as e:
        print(e)
        raise

    if submissions:
        create_comments_excel_file(survey_id, period, submissions)
    else:
        print("No submissions, exiting script")
